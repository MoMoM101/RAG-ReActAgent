import csv
import logging
from collections.abc import Iterable, Sequence
from itertools import islice
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

_MAX_TABULAR_DATA_ROWS = 10_000


def _markdown_cell(value: Any) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .replace("\r\n", "<br>")
        .replace("\r", "<br>")
        .replace("\n", "<br>")
        .replace("|", r"\|")
    )


def _rows_to_markdown(rows: Iterable[Sequence[Any]]) -> str:
    """Render a bounded row iterator as a deterministic Markdown table."""
    iterator = iter(rows)
    try:
        header = list(next(iterator))
    except StopIteration:
        return ""

    data_rows = [list(row) for row in islice(iterator, _MAX_TABULAR_DATA_ROWS)]
    column_count = max([len(header), *(len(row) for row in data_rows)])
    if column_count == 0:
        return ""

    header.extend("" for _ in range(column_count - len(header)))
    lines = [
        "| " + " | ".join(_markdown_cell(value) for value in header) + " |",
        "| " + " | ".join("---" for _ in range(column_count)) + " |",
    ]
    for row in data_rows:
        row.extend("" for _ in range(column_count - len(row)))
        lines.append("| " + " | ".join(_markdown_cell(value) for value in row) + " |")
    return "\n".join(lines)


def load_pdf(file_path: str) -> str:
    import fitz  # PyMuPDF

    from config import settings

    doc = fitz.open(file_path)
    text_parts = []
    for page in doc:
        text = page.get_text()

        # OCR embedded images on the page (covers mixed text+image pages)
        ocr_texts: list[str] = []
        if settings.ocr_enabled:
            try:
                from ocr.factory import create_ocr
                from ocr.preprocess import image_from_bytes
                ocr = create_ocr()
                if ocr:
                    for img_info in page.get_images(full=True):
                        xref = img_info[0]
                        base_image = doc.extract_image(xref)
                        img_bytes = base_image["image"]
                        try:
                            preprocessed = image_from_bytes(img_bytes, dpi=200)
                            ocr_text = ocr.recognize(preprocessed)
                        except Exception:
                            try:
                                ocr_text = ocr.recognize_from_bytes(img_bytes, dpi=200)
                            except Exception:
                                logger.warning("OCR failed both preprocess and raw, embedded image size=%d", len(img_bytes))
                                continue
                        if ocr_text.strip():
                            ocr_texts.append(ocr_text)
            except Exception:
                pass  # OCR not critical, continue with extracted text

        # If extracted text is too sparse, OCR the whole page as fallback (scanned PDF)
        if settings.ocr_enabled and len(text.strip()) < settings.ocr_min_text_length:
            try:
                if not ocr_texts:
                    from ocr.factory import create_ocr
                    from ocr.preprocess import image_from_bytes
                    ocr = create_ocr()
                    if ocr:
                        pix = page.get_pixmap(dpi=200)
                        img_bytes = pix.tobytes("png")
                        try:
                            preprocessed = image_from_bytes(img_bytes, dpi=200)
                            text = ocr.recognize(preprocessed)
                        except Exception:
                            try:
                                text = ocr.recognize_from_bytes(img_bytes, dpi=200)
                            except Exception:
                                logger.warning("OCR failed both preprocess and raw, full page size=%d", len(img_bytes))
            except Exception:
                pass

        if ocr_texts:
            text += "\n" + "\n".join(ocr_texts)

        text_parts.append(text)
    doc.close()
    return "\n\n".join(text_parts)


def load_docx(file_path: str) -> str:
    from docx import Document
    doc = Document(file_path)
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def load_txt(file_path: str) -> str:
    return Path(file_path).read_text(encoding="utf-8")


def load_md(file_path: str) -> str:
    return Path(file_path).read_text(encoding="utf-8")


def load_csv(file_path: str) -> str:
    with open(file_path, encoding="utf-8-sig", newline="") as file:
        return _rows_to_markdown(csv.reader(file))


def load_xlsx(file_path: str) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return _rows_to_markdown(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()


def load_image(file_path: str) -> str:
    from config import settings

    with open(file_path, "rb") as f:
        data = f.read()
    if not settings.ocr_enabled:
        return ""
    try:
        from ocr.factory import create_ocr
        from ocr.preprocess import image_from_bytes
        ocr = create_ocr()
        if ocr is None:
            return ""
        try:
            preprocessed = image_from_bytes(data, dpi=200)
            return ocr.recognize(preprocessed)
        except Exception:
            try:
                return ocr.recognize_from_bytes(data)
            except Exception:
                logger.warning("OCR failed both preprocess and raw, image size=%d", len(data))
                return ""
    except Exception:
        return ""


LOADERS = {
    ".pdf": load_pdf,
    ".docx": load_docx,
    ".txt": load_txt,
    ".md": load_md,
    ".csv": load_csv,
    ".xlsx": load_xlsx,
    ".jpg": load_image,
    ".jpeg": load_image,
    ".png": load_image,
}


def load_document(file_path: str, file_type: str) -> str:
    loader = LOADERS.get(file_type.lower())
    if not loader:
        raise ValueError(f"Unsupported file type: {file_type}")
    return loader(file_path)
